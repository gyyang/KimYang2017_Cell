# -*- coding: utf-8 -*-
"""
A replication of Chaudhuri et al. Neuron 2015

@author: Guangyu Robert Yang, 2015/11
"""
from __future__ import division

import pickle
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

#---------------------------------------------------------------------------------
# Process Allen Connectivity Data
#---------------------------------------------------------------------------------

def get_area_names():
    # Get areas' names, full names, and major regions
    wb0 = load_workbook('library/nature13186-s2.xlsx')
    sheet = wb0.get_sheet_by_name('Voxel Count_295 Structures')
    acronym_list        = list()
    full_name_list      = list()
    major_region_list   = list()
    
    for i in range(1,400):
        acronym = sheet.cell(row=i, column=2).value
        if acronym is None:
            break
        acronym_list.append(str(acronym))
        full_name_list.append(str(sheet.cell(row=i, column=3).value))
        major_region_list.append(str(sheet.cell(row=i, column=4).value))
    
    data = {'acronym_list'      : acronym_list,
            'full_name_list'    : full_name_list,
            'major_region_list' : major_region_list}
    with open('library/mouse_area_names_allen.pkl','wb') as f:
        pickle.dump(data,f)


def get_connectivity():
    # Get their corresponding connection weights

    wb = load_workbook('library/nature13186-s4.xlsx')
    '''
    Quantitative projection strength values underlying the linear model based
    Connectivity Matrix in figure 4a. The 213 anatomical regions as both
    source regions (in rows) and target regions on both ipsilateral and contralateral
    hemispheres (in columns) are shown here. See Supplementary Table 1 for the
    corresponding full name and acronym of each region.
    '''

    data = dict()

    for sheet_name in ['W_ipsi', 'W_contra']:
        W_sheet = wb.get_sheet_by_name(sheet_name)

        areas0 = list()
        for i in range(1,300):
            name = W_sheet.cell(row=0, column=i).value
            if name is None:
                break
            areas0.append(str(name))

        n_area = len(areas0)

        W0 = np.zeros((n_area,n_area))
        for i in range(n_area):
            for j in range(n_area):
                W0[i,j] = W_sheet.cell(row=j+1,column=i+1).value
                if W0[i,j] is None:
                    print 'Error reading W'

        data[sheet_name] = W0

    data['areas']   = areas0
    data['W_mean']  = (data['W_ipsi'] + data['W_contra'])/2
    
    with open('library/mouse_connectivity_data_allen.pkl','wb') as f:
        pickle.dump(data,f)


def get_data_allen():
    
    with open('library/mouse_area_names.pkl','rb') as f:
        area_data = pickle.load(f)
    
    with open('library/mouse_connectivity_data_allen.pkl','rb') as f:
        W_data = pickle.load(f)
            
    area_idx_list   = list()
    areas           = list()
    full_names      = list()
    
    for i, area in enumerate(W_data['areas']):
        j = area_data['acronym_list'].index(area)
        if area_data['major_region_list'][j] == 'Isocortex' and area != 'FRP':
            area_idx_list.append(i)
            areas.append(area)
            full_names.append(area_data['full_name_list'][j])
    
    W_ipsi      = W_data['W_ipsi'][area_idx_list,:][:,area_idx_list]
    W_contra    = W_data['W_contra'][area_idx_list,:][:,area_idx_list]
    W_mean      = W_data['W_mean'][area_idx_list,:][:,area_idx_list]
    
    data        = {'areas'          : areas,
                   'full_names'     : full_names,
                   'W_ipsi'         : W_ipsi,
                   'W_contra'       : W_contra,
                   'W_mean'         : W_mean,
                   'dataset'        : 'allen'}
    
    with open('library/mouse_isocortex_data_allen.pkl','wb') as f:
        pickle.dump(data,f)



with open('library/mouse_area_names.pkl','rb') as f:
    area_data = pickle.load(f)

with open('library/mouse_connectivity_data_allen.pkl','rb') as f:
    W_data = pickle.load(f)
        
area_idx_list   = list()
areas           = list()
full_names      = list()

for i, area in enumerate(W_data['areas']):
    j = area_data['acronym_list'].index(area)
    if area_data['major_region_list'][j] == 'Isocortex' and area != 'FRP':
        area_idx_list.append(i)
        areas.append(area)
        full_names.append(area_data['full_name_list'][j])

W_ipsi      = W_data['W_ipsi'][area_idx_list,:][:,area_idx_list]
W_contra    = W_data['W_contra'][area_idx_list,:][:,area_idx_list]
W_mean      = W_data['W_mean'][area_idx_list,:][:,area_idx_list]

data        = {'areas'          : areas,
               'full_names'     : full_names,
               'W_ipsi'         : W_ipsi,
               'W_contra'       : W_contra,
               'W_mean'         : W_mean,
               'dataset'        : 'allen'}
#get_data_allen()

#---------------------------------------------------------------------------------
# Process USC Connectivity Data
#---------------------------------------------------------------------------------

def get_data_usc():
    # Load Zingg results, and convert into the Allen Brain Institute data form
    W_retro  = np.loadtxt('library/Zingg2014/zingg2014_Cweighted_retrograde.txt')
    W_antero = np.loadtxt('library/Zingg2014/zingg2014_Cweighted_anterograde.txt')

    # Load USC data
    with open('library/Zingg2014/zingg2014_areas.txt','rb') as f:
        areas_usc = f.read().splitlines()

    # Load Allen Brain Institute data
    with open('library/mouse_isocortex_data.pkl','rb') as f:
        p_allen = pickle.load(f)

    # Load Kim data
    with open('library/mouse_interneuron_density.pkl','rb') as f:
        p_kim = pickle.load(f)

    areas_kim = p_kim['areas']

    areas_allen = p_allen['areas']

    # Start conversion
    areas = areas_allen
    areas.remove('FRP') # take out areas not in USC dataset
    areas.remove('VISpl')

    def convert1d_usc2allen(areas,W0,W1):
        #TODO: Notice below is a very rough draft conversion
        for i, area in enumerate(areas):
            if area == 'MOp':
                idx_list = [areas_usc.index(a) for a in ['MOp-orf','MOp-ll \& tr','MOp-w','MOp-ul']]
            elif area == 'MOs':
                idx_list = [areas_usc.index(a) for a in ['MOs-c','MOs-fef','MOs-rd','MOs-rdl','MOs-rdm']]
            elif area == 'SSp-bfd':
                idx_list = [areas_usc.index(a) for a in ['SSp-bfd.al','SSp-bfd.cm']]
            elif area == 'SSs':
                idx_list = [areas_usc.index(a) for a in ['SSs-cd','SSs-r \& cv']]
            elif area in ['SSp-ll','SSp-tr']:
                idx_list = [areas_usc.index('SSp-ll \& tr')]
            elif area in ['SSp-m','SSp-n']:
                idx_list = [areas_usc.index('SSp-m \& n')]
            else:
                idx_list = [areas_usc.index(area)]
            W1[i] = W0[idx_list].mean(axis=0)
        return W1

    W0 = (W_retro + W_antero)/2
    W1 = np.zeros((len(areas),len(areas_usc)))
    W1 = convert1d_usc2allen(areas,W0,W1)
    W0 = W1.T
    W1 = np.zeros((len(areas),len(areas)))
    W1 = convert1d_usc2allen(areas,W0,W1)
    W  = W1.T

    data        = {'areas'          : areas,
                   'W_mean'         : W,
                   'dataset'        : 'usc'}

    with open('library/mouse_isocortex_data_usc.pkl','wb') as f:
            pickle.dump(data,f)
            

#---------------------------------------------------------------------------------
# Process Interneuron Density Data
#---------------------------------------------------------------------------------

def get_interneuron_density():
    wb0 = load_workbook('library/Gad2_counting.xlsx')
    sheet = wb0.get_sheet_by_name('Sheet1')
    acronym_list0  = list()
    gad_list0      = list()
    for i in range(1,400):
        acronym = sheet.cell(row=i, column=0).value
        if acronym is None:
            break
        acronym_list0.append(str(acronym))
        gad_list0.append(sheet.cell(row=i, column=4).value)
    
    
    # Get areas' names, full names, and major regions
    wb1 = load_workbook('library/Density_interneuron_cortical.xlsx')
    sheet = wb1.get_sheet_by_name('by layers_isocortex')
    acronym_list = list()
    pv_list      = list()
    sst_list     = list()
    vip_list     = list()
    gad_list     = list()
    
    for i in range(2,400):
        acronym = sheet.cell(row=i, column=0).value
        if acronym is None:
            break
        acronym_list.append(str(acronym))
        pv_list.append(sheet.cell(row=i, column=1).value)
        sst_list.append(sheet.cell(row=i, column=2).value)
        vip_list.append(sheet.cell(row=i, column=3).value)
        gad_list.append(gad_list0[acronym_list0.index(acronym)])
    
    data = {'areas'    : acronym_list,
            'pv_list'  : np.array(pv_list),
            'sst_list' : np.array(sst_list),
            'vip_list' : np.array(vip_list),
            'gad_list' : np.array(gad_list)}
            
    with open('library/mouse_interneuron_density.pkl','wb') as f:
        pickle.dump(data,f)
        
#get_interneuron_density()
        
        
def get_interneuron_density_flatmap():
    # Get areas' names, full names, and major regions
    wb0 = load_workbook('library/ML-corticalFlatmap_box_heatmap.xlsx')
    data = dict()    
    for name in ['PV','SST','VIP','Gad2']:
        sheet = wb0.get_sheet_by_name(name)
        ml_list = list() # medial-lateral coordinate
        ap_list = list() # anterior-posterior coordinate
        density_list = list() # density
        
        for i in range(5,466):
            ml_list.append(sheet.cell(row=i, column=0).value)
            ap_list.append(sheet.cell(row=i, column=1).value)
            density_list.append(sheet.cell(row=i, column=2).value)
        
        data[name] = {'ml_list'      : np.array(ml_list),
                    'ap_list'        : np.array(ap_list),
                    'density_list'   : np.array(density_list)}
                    
    with open('library/mouse_interneuron_density_flatmap.pkl','wb') as f:
        pickle.dump(data,f)
        
#get_interneuron_density_flatmap()
        
#---------------------------------------------------------------------------------
# Miscellaneous
#---------------------------------------------------------------------------------

def zingg_areadivision():
    # See Zingg et al.
    div = dict()
    div['somatic'] = ['MOp', 'MOs', 'SSp-bfd', 'SSp-ll', 'SSp-m',
                     'SSp-n', 'SSp-tr', 'SSp-ul', 'SSs']
    
    div['medial'] = [ 'VISal', 'VISam', 'VISl', 'VISp','VISpm', 'ORBl',
                      'ORBvl', 'ACAd', 'ACAv', 'RSPagl', 'RSPd', 'RSPv',
                      'AUDd', 'AUDp', 'AUDv','PTLp'] + ['VISpl','AUDpo']

    div['vis-aud'] = [ 'VISal', 'VISam', 'VISl', 'VISp','VISpm',
                      'AUDd', 'AUDp', 'AUDv'] + ['VISpl','AUDpo']

    div['medial associa.'] = [ 'ORBl', 'ORBvl', 'ACAd', 'ACAv', 'RSPagl', 'RSPd', 'RSPv','PTLp']
    
    div['lateral'] = ['VISC','GU',  'AId', 'AIp', 'AIv', 'TEa','PERI', 'ECT']
    
    div['medial prefrontal'] = ['ORBm','ILA','PL']
    
    div_map = dict()
    for div_name in ['somatic','medial','lateral','medial prefrontal']:
        for area in div[div_name]:
            div_map[area] = div_name
            
    with open('library/Zingg_areadivision.pkl','wb') as f:
        pickle.dump(div,f)

#zingg_areadivision()